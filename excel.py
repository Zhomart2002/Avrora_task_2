import enum
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl import Workbook

border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
    

def create_headers(sheet):
    font = Font(bold = True, size=11)
    fill = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
    alignment = Alignment(horizontal='center', vertical='center')

    headers = [['Наименование организации', 40], # Name of header and its column size
                ['БИН организации', 20],
                ['ФИО руководителя', 30],
                ['ИИН руководителя', 20],
                ['Полный адрес организации', 50]]

    for index, column_letter in enumerate('DEFGH'):
        cell_address = f"{column_letter}4"
        sheet[cell_address].border = border
        sheet[cell_address].font = font
        sheet[cell_address].fill = fill
        sheet[cell_address].alignment = alignment
        sheet[cell_address] = headers[index][0]
        sheet.column_dimensions[column_letter].width = headers[index][1]

    sheet.row_dimensions[4].height = 20
    
def set_data(sheet, data):
    current_row = 5

    for organization_data in data:
        for index, col in enumerate('DEFGH'):
            cell_address = f"{col}{current_row}"
            sheet[cell_address] = organization_data[index]
            sheet[cell_address].border = border
        current_row += 1

def export_to_excel(data): # [[Наименование организации, БИН организации, ФИО руководителя, ИИН руководителя, Полный адрес организации], ...]
    
    work_book = Workbook()
    sheet = work_book.active
    sheet.title = "Result"

    create_headers(sheet)
    set_data(sheet, data)

    work_book.save("Result.xlsx")
    work_book.close()

if __name__ == '__main__':
    export_to_excel()